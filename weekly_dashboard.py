# -*- coding: utf-8 -*-
"""
睡sleepp的店 · 每周经营看板生成器（磨砂玻璃·四段式布局版）
- 自动识别累计订单数据 / 指定 xlsx（千帆导出）
- 清洗 -> 计算核心指标 -> 渲染自包含 HTML 看板 + Markdown 报告 + 清洗 CSV + Excel 报表
- 风格：低饱和磨砂玻璃质感 + 四段式信息层级（宏观大盘→订单结构→固定报表→销售热点）
- 新增：下单时段/周几分布、月度/季度可复制报表、Excel 导出
- 用法: python weekly_dashboard.py  [可选: 指定xlsx路径]
"""
import pandas as pd, numpy as np, json, glob, os, sys, openpyxl, warnings
from datetime import datetime
from collections import Counter

WS = r"D:\ai\小红书引流笔记工作流/数据记录与分析"
amt_col = "商家应收金额(元)（支付金额）"
pay_col = "用户应付金额(元)"

# ── KPOP 团体关键词映射 ──
GROUP_MAP = [
    ("ILLIT",       ["illit"]),
    ("ITZY",        ["itzy"]),
    ("Nmixx",       ["nmixx"]),
    ("Kiiikiii",    ["kiiikiii", "稞"]),
    ("Cortis",      ["cortis"]),
    ("LE SSERAFIM", ["lesserafim", "lef", "炽"]),
    ("aespa",       ["aespa"]),
    ("NewJeans",    ["newjeans", "new jeans"]),
    ("Babymonster", ["babymonster", "baby monster"]),
    ("Stray Kids",  ["stray kids", "skz"]),
    ("SEVENTEEN",   ["seventeen", "svt"]),
    ("ENHYPEN",     ["enhypen"]),
    ("RIIZE",       ["riize"]),
    ("TWS",         ["tws"]),
    ("ZB1",         ["zerobaseone", "zb1"]),
    ("Kiss of Life",["kiss of life", "kiof"]),
]

PRODUCT_TYPE_MAP = [
    ("专辑",   ["专辑", "album"]),
    ("小卡",   ["小卡", "photocard", "pc"]),
    ("官娃",   ["官娃", "娃娃", "doll"]),
    ("摆件",   ["摆件", "陶瓷", "手办"]),
    ("海报",   ["海报", "poster"]),
    ("徽章",   ["徽章", "badge"]),
    ("亚克力", ["亚克力", "acrylic"]),
    ("钥匙扣", ["钥匙扣", "keyring"]),
    ("写真",   ["写真", "photobook"]),
    ("盲盒",   ["盲盒", "blind"]),
]

# 时段定义
TIME_PERIODS = [
    ("凌晨", 0, 5),
    ("上午", 6, 12),
    ("下午", 13, 17),
    ("晚上", 18, 22),
    ("深夜", 23, 23),
]

WEEKDAY_NAMES = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]


def group_of(name):
    s = str(name).lower()
    for g, kws in GROUP_MAP:
        for kw in kws:
            if kw in s:
                return g
    return "其他/未分类"


def extract_keywords(name):
    s = str(name)
    sl = s.lower()
    res = []
    g = group_of(s)
    if g != "其他/未分类":
        res.append(g)
    for label, kws in PRODUCT_TYPE_MAP:
        for kw in kws:
            if kw.lower() in sl:
                res.append(label)
                break
    return res


def read_orders(path):
    """稳健读取包裹详情：pandas 优先，空表时回退 openpyxl 强制全量迭代。"""
    warnings.filterwarnings("ignore")
    df = None
    try:
        df = pd.read_excel(path, sheet_name="包裹详情", engine="openpyxl")
        if df.shape[1] < 2 or df.shape[0] == 0:
            df = None
    except Exception:
        df = None
    if df is None:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb["包裹详情"]
        rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True))
        wb.close()
        df = pd.DataFrame(rows[1:], columns=rows[0])
    for c in [amt_col, pay_col, "SKU件数"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")
    return df


def _color(metric, v):
    if metric == "取消率":
        return "#C47070" if v >= 20 else ("#D4A050" if v >= 10 else "#7AAA90")
    if metric == "售后率":
        return "#C47070" if v >= 10 else ("#D4A050" if v >= 5 else "#7AAA90")
    if metric == "复购率":
        return "#C47070" if v < 10 else ("#D4A050" if v < 25 else "#7AAA90")
    return "#5A8A78"


def _export_excel(R, paid_df, df):
    """导出月度/季度汇总 Excel 报表（可直接复制到 Excel 的格式）"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows

    out = os.path.join(WS, "report_%s.xlsx" % datetime.now().strftime("%Y%m%d"))
    wb = Workbook()

    # ── Sheet 1: 月度报表 ──
    ws = wb.active
    ws.title = "月度汇总"
    thin = Side(style='thin', color='FFD0DCD4')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(start_color="E8F2EE", end_color="E8F2EE", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FF3E5A4E")
    data_font = Font(size=10, color="FF4A5E54")

    # 写标题
    ws['A1'] = "睡sleepp的店 · 月度经营报表"
    ws['A1'].font = Font(bold=True, size=14, color="FF3E5A4E")
    ws.merge_cells('A1:I1')
    ws['A2'] = "数据周期：%s  |  生成：%s" % (R["数据周期"], R["生成时间"])
    ws['A2'].font = Font(size=10, color="FF889990")
    ws.merge_cells('A2:I2')

    headers = ["月份", "订单数", "净GMV(¥)", "客单价(¥)", "付费用户", "取消率%", "售后率%", "复购率%", "预售占比%"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    monthly = R["月度报表"]
    for ri, m in enumerate(monthly, 5):
        vals = [m["月份"], m["订单数"], m["GMV"], m["客单价"], m["用户数"],
                m["取消率"], m["售后率"], m["复购率"], m["预售占比"]]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.font = data_font
            cell.border = border
            if ci >= 2:
                cell.alignment = Alignment(horizontal='right')

    # 列宽
    widths = [10, 10, 14, 12, 12, 10, 10, 10, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64+i)].width = w

    # ── Sheet 2: 季度报表 ──
    ws2 = wb.create_sheet("季度汇总")
    ws2['A1'] = "睡sleepp的店 · 季度经营报表"
    ws2['A1'].font = Font(bold=True, size=14, color="FF3E5A4E")
    ws2.merge_cells('A1:I1')
    ws2['A2'] = "数据周期：%s  |  生成：%s" % (R["数据周期"], R["生成时间"])
    ws2['A2'].font = Font(size=10, color="FF889990")
    ws2.merge_cells('A2:I2')

    q_headers = ["季度", "订单数", "净GMV(¥)", "客单价(¥)", "付费用户", "取消率%", "售后率%", "复购率%", "预售占比%"]
    for ci, h in enumerate(q_headers, 1):
        cell = ws2.cell(row=4, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    quarterly = R["季度报表"]
    for ri, q in enumerate(quarterly, 5):
        vals = [q["季度"], q["订单数"], q["GMV"], q["客单价"], q["用户数"],
                q["取消率"], q["售后率"], q["复购率"], q["预售占比"]]
        for ci, v in enumerate(vals, 1):
            cell = ws2.cell(row=ri, column=ci, value=v)
            cell.font = data_font
            cell.border = border
            if ci >= 2:
                cell.alignment = Alignment(horizontal='right')

    for i, w in enumerate(widths, 1):
        ws2.column_dimensions[chr(64+i)].width = w

    # ── Sheet 3: 下单时段分布 ──
    ws3 = wb.create_sheet("下单时段分布")
    ws3['A1'] = "用户下单习惯分析"
    ws3['A1'].font = Font(bold=True, size=14, color="FF3E5A4E")
    ws3.merge_cells('A1:E1')

    t_headers = ["时段", "订单数", "GMV(¥)", "客单价(¥)", "占比%"]
    for ci, h in enumerate(t_headers, 1):
        cell = ws3.cell(row=3, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    for ri, t in enumerate(R["时段分布"], 4):
        vals = [t["时段"], t["订单数"], t["GMV"], t["客单价"], t["占比"]]
        for ci, v in enumerate(vals, 1):
            cell = ws3.cell(row=ri, column=ci, value=v)
            cell.font = data_font
            cell.border = border

    ws3.cell(row=3, column=7, value="周几分布").font = header_font
    wd_headers = ["周几", "订单数", "GMV(¥)", "客单价(¥)", "占比%"]
    for ci, h in enumerate(wd_headers, 7):
        cell = ws3.cell(row=3, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    for ri, w in enumerate(R["周几分布"], 4):
        vals = [w["周几"], w["订单数"], w["GMV"], w["客单价"], w["占比"]]
        for ci, v in enumerate(vals, 7):
            cell = ws3.cell(row=ri, column=ci, value=v)
            cell.font = data_font
            cell.border = border

    wb.save(out)
    print("Excel 报表 ->", os.path.basename(out))
    return out


# ══════════════════════════════════════
#  HTML 渲染（磨砂玻璃 · 四段式布局）
# ══════════════════════════════════════
def build_html(R):
    def card(label, val, sub="", color="#5A8A78"):
        return ('<div class="card"><div class="cl">%s</div><div class="cv" style="color:%s">%s</div>'
                '<div class="cs">%s</div></div>' % (label, color, val, sub))

    cards = []
    cards.append(card("净GMV", "¥%.0f" % R["净GMV"], "已完成 ¥%.0f" % R["已完成GMV"]))
    cards.append(card("订单数", "%d" % R["订单数"], "人均 %.2f 单" % R["人均订单"]))
    cards.append(card("付费用户", "%d" % R["用户数"], ""))
    cards.append(card("客单价", "¥%.0f" % R["客单价"], "中位 ¥%.0f" % R["客单价中位"]))
    cards.append(card("取消率", "%.1f%%" % R["取消率"], "健康<10%", _color("取消率", R["取消率"])))
    cards.append(card("售后率", "%.1f%%" % R["售后率"], "健康<5%", _color("售后率", R["售后率"])))
    cards.append(card("复购率", "%.1f%%" % R["复购率"], "目标>25%", _color("复购率", R["复购率"])))
    cards.append(card("预售占比", "%.0f%%" % R["预售占比"], ""))
    cards_html = "".join(cards)

    # ── 表格构建函数 ──
    def sku_table():
        rows = ""
        for i, r in enumerate(R["TOP_SKU"], 1):
            rows += "<tr><td>%d</td><td>%s</td><td>¥%d</td><td>%d</td></tr>" % (i, r["name"], r["gmv"], r["orders"])
        return '<table><thead><tr><th>#</th><th>SKU</th><th>GMV</th><th>订单</th></tr></thead><tbody>%s</tbody></table>' % rows

    def prov_table():
        rows = ""
        for i, r in enumerate(R["省份"], 1):
            rows += "<tr><td>%d</td><td>%s</td><td>¥%d</td><td>%d</td><td>%d</td></tr>" % (i, r["name"], r["gmv"], r["orders"], r["users"])
        return '<table><thead><tr><th>#</th><th>省份</th><th>GMV</th><th>订单</th><th>用户</th></tr></thead><tbody>%s</tbody></table>' % rows

    def group_table():
        rows = ""
        for i, r in enumerate(R["团体"], 1):
            rows += "<tr><td>%d</td><td>%s</td><td>¥%d</td><td>%d</td></tr>" % (i, r["name"], r["gmv"], r["orders"])
        return '<table><thead><tr><th>#</th><th>团体</th><th>净GMV</th><th>订单</th></tr></thead><tbody>%s</tbody></table>' % rows

    def spotpre_table():
        sp, pr = R["现货"], R["预售"]
        return ('<table><thead><tr><th>指标</th><th>现货</th><th>预售</th></tr></thead><tbody>'
                '<tr><td>订单数</td><td>%d</td><td>%d</td></tr>'
                '<tr><td>净GMV</td><td>¥%d</td><td>¥%d</td></tr>'
                '<tr><td>客单价</td><td>¥%.0f</td><td>¥%.0f</td></tr>'
                '<tr><td>取消率</td><td>%.1f%%</td><td>%.1f%%</td></tr>'
                '</tbody></table>' % (sp["订单数"], pr["订单数"], sp["GMV"], pr["GMV"],
                                       sp["客单价"], pr["客单价"], sp["取消率"], pr["取消率"]))

    def month_report_table():
        """月度报表——可选中复制的表格"""
        rows = ""
        for m in R["月度报表"]:
            rows += ('<tr><td>' + m["月份"] + '</td><td class="num">' + str(m["订单数"]) + '</td><td class="num">¥' + str(m["GMV"]) +
                     '</td><td class="num">¥' + str(round(m["客单价"])) + '</td><td class="num">' + str(m["用户数"]) +
                     '</td><td class="num">' + str(m["取消率"]) + '</td><td class="num">' + str(m["售后率"]) +
                     '</td><td class="num">' + str(m["复购率"]) + '</td><td class="num">' + str(m["预售占比"]) + '</td></tr>')
        return ('<table class="rpt-table"><thead><tr><th>月份</th><th>订单数</th><th>净GMV</th>'
                '<th>客单价</th><th>付费用户</th><th>取消率%</th><th>售后率%</th>'
                '<th>复购率%</th><th>预售占比%</th></tr></thead><tbody>' + rows + '</tbody></table>')

    def quarter_report_table():
        rows = ""
        for q in R["季度报表"]:
            rows += ('<tr><td>' + q["季度"] + '</td><td class="num">' + str(q["订单数"]) + '</td><td class="num">¥' + str(q["GMV"]) +
                     '</td><td class="num">¥' + str(round(q["客单价"])) + '</td><td class="num">' + str(q["用户数"]) +
                     '</td><td class="num">' + str(q["取消率"]) + '</td><td class="num">' + str(q["售后率"]) +
                     '</td><td class="num">' + str(q["复购率"]) + '</td><td class="num">' + str(q["预售占比"]) + '</td></tr>')
        return ('<table class="rpt-table"><thead><tr><th>季度</th><th>订单数</th><th>净GMV</th>'
                '<th>客单价</th><th>付费用户</th><th>取消率%</th><th>售后率%</th>'
                '<th>复购率%</th><th>预售占比%</th></tr></thead><tbody>' + rows + '</tbody></table>')

    def time_dist_bars():
        """时段分布横向条形图（纯 CSS 实现）"""
        max_gmv = max((t["GMV"] for t in R["时段分布"]), default=1)
        bars = ""
        for t in R["时段分布"]:
            pct = t["GMV"] / max_gmv * 100 if max_gmv > 0 else 0
            bars += ('<div class="dist-row"><div class="dist-label">%s</div>'
                     '<div class="dist-bar-wrap"><div class="dist-bar" style="width:%.1f%%"></div></div>'
                     '<div class="dist-val">%d单 ¥%d</div></div>' % (t["时段"], pct, t["订单数"], t["GMV"]))
        return bars

    def weekday_dist_bars():
        """周几分布"""
        max_gmv = max((w["GMV"] for w in R["周几分布"]), default=1)
        bars = ""
        for w in R["周几分布"]:
            pct = w["GMV"] / max_gmv * 100 if max_gmv > 0 else 0
            bars += ('<div class="dist-row"><div class="dist-label">%s</div>'
                     '<div class="dist-bar-wrap"><div class="dist-bar wd-bar" style="width:%.1f%%"></div></div>'
                     '<div class="dist-val">%d单 ¥%d</div></div>' % (w["周几"], pct, w["订单数"], w["GMV"]))
        return bars

    def week_hot_panel():
        chips = "".join('<span class="chip">%s <b>%d</b></span>' % (k["word"], k["count"]) for k in R["本周热词"])
        rows = ""
        for i, r in enumerate(R["本周TOP"], 1):
            rows += "<tr><td>%d</td><td>%s</td><td>¥%d</td><td>%d</td></tr>" % (i, r["name"], r["gmv"], r["orders"])
        return ('<div class="kw-chips">%s</div>'
                '<div class="note" style="margin:10px 0 6px">本周 TOP 热销 SKU</div>'
                '<table><thead><tr><th>#</th><th>SKU</th><th>GMV</th><th>订单</th></tr></thead><tbody>%s</tbody></table>'
                % (chips, rows))

    def last_week_panel():
        su = R["上周汇总"]
        sum_chips = ('<span class="wk-chip">共 %d 单</span><span class="wk-chip">现货 %d</span>'
                     '<span class="wk-chip">预售 %d</span><span class="wk-chip canc">已取消 %d</span>'
                     '<span class="wk-chip ok">净GMV ¥%d</span>'
                     % (su["单数"], su["现货"], su["预售"], su["取消"], su["净GMV"]))
        return sum_chips, R["上周订单html"]

    # ── 警告 & 元信息 ──
    fresh = ('<div class="warn">⚠ 数据已 %d 天未更新（最近订单 %s），请导出最新千帆订单后重跑看板</div>'
             % (R["距今天数"], R["最近订单日"])) if R["距今天数"] > 7 else ""
    freshword = "数据新鲜" if R["距今天数"] <= 7 else "数据偏旧"
    wow = ""
    if R["周环比"] is not None:
        wow = "上周GMV ¥%d，环比 %s%%" % (R["最近周GMV"], ("+%.1f" % R["周环比"]) if R["周环比"] >= 0 else ("%.1f" % R["周环比"]))
    elif R.get("最近周GMV"):
        wow = "上周GMV ¥%d" % R["最近周GMV"]
    pay_str = "、".join("%s %d" % (k, v) for k, v in R["支付方式"].items())
    exp_str = "、".join("%s %d" % (k, v) for k, v in R["快递"].items())

    # 收集所有面板内容
    wk_sum_html, wk_tbl_html = last_week_panel()

    tpl = """<!DOCTYPE html><html lang="zh-CN"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>睡sleepp的店 · 经营看板</title>
<style>
  /* ── 全局：磨砂玻璃基底 ── */
  :root{
    --glass-bg:rgba(255,255,255,.55);
    --glass-border:rgba(180,200,190,.30);
    --glass-shadow:0 8px 32px rgba(100,140,120,.10);
    --text-primary:#2D3E36;
    --text-secondary:#6B8076;
    --accent:#5A9A7E;
    --accent-light:#A8D4BC;
    --danger-bg:rgba(196,112,112,.12);
    --danger-text:#B06A6A;
    --warn-bg:rgba(212,160,80,.12);
    --warn-text:#A08840;
    --ok-bg:rgba(122,170,144,.12);
    --ok-text:#4A8A68;
    --section-gap:32px;
  }
  *{box-sizing:border-box;}
  body{font-family:-apple-system,'Segoe UI','Microsoft YaHei','Noto Sans SC',sans-serif;
    margin:0;padding:0;min-height:100vh;
    /* 极低饱和绿灰渐变底色（比参考图更淡更灰） */
    background:linear-gradient(160deg,#EDF5F0 0%,#E4F0EA 25%,#DAEBE3 50%,#D0E5DB 75%,#C8DFD6 100%);
    color:var(--text-primary);line-height:1.5;
    -webkit-font-smoothing:antialiased;}
  /* 磨砂噪点纹理层 */
  body::before{content:"";position:fixed;inset:0;z-index:0;pointer-events:none;
    opacity:.035;
    background-image:url("data:image/svg+xml,%3Csvg viewBox='0 0 256 256' xmlns='http://www.w3.org/2000/svg'%3E%3Cfilter id='n'%3E%3CfeTurbulence type='fractalNoise' baseFrequency='.85' numOctaves='4' stitchTiles='stitch'/%3E%3C/filter%3E%3Crect width='100%25' height='100%25' filter='url(%23n)'/%3E%3C/svg%3E");
    background-size:180px 180px;}
  /* 光斑层 */
  body::after{content:"";position:fixed;inset:0;z-index:0;pointer-events:none;
    background:radial-gradient(ellipse 600px 400px at 20% 10%, rgba(168,212,188,.18), transparent),
              radial-gradient(ellipse 500px 350px at 80% 70%, rgba(200,222,210,.14), transparent);}

  .wrap{position:relative;z-index:1;max-width:1160px;margin:0 auto;padding:28px 24px 48px;}

  /* ── 标题区 ── */
  header{margin-bottom:28px;}
  h1{font-size:22px;font-weight:700;margin:0 0 6px;color:var(--text-primary);letter-spacing:.5px;}
  h1 .spark{color:var(--accent-light);font-size:16px;margin:0 4px;}
  .meta{font-size:12px;color:var(--text-secondary);}
  .wow-badge{display:inline-block;background:var(--glass-bg);backdrop-filter:blur(12px);
    border:1px solid var(--glass-border);border-radius:999px;padding:4px 14px;font-size:12px;
    color:var(--text-secondary);margin-top:6px;}

  .warn{background:var(--danger-bg);border:1px solid rgba(196,112,112,.25);color:var(--danger-text);
    padding:10px 16px;border-radius:14px;font-size:13px;margin-bottom:16px;
    backdrop-filter:blur(8px);}

  /* ── Section 容器 ── */
  .sec{margin-bottom:var(--section-gap);}
  .sec-title{font-size:13px;font-weight:600;color:var(--text-secondary);text-transform:uppercase;
    letter-spacing:2px;margin-bottom:14px;padding-bottom:8px;
    border-bottom:1px solid rgba(180,200,190,.25;display:flex;align-items:center;gap:8px;}
  .sec-title .sec-num{display:inline-flex;align-items:center;justify-content:center;width:22px;height:22px;
    border-radius:8px;background:var(--accent-light);color:#fff;font-size:11px;font-weight:700;}
  .sec-title .sec-name{color:var(--text-primary);font-size:15px;text-transform:none;letter-spacing:1px;}

  /* ── 卡片（KPI） ── */
  .cards{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;}
  .card{background:var(--glass-bg);backdrop-filter:blur(18px);
    border:1px solid var(--glass-border);border-radius:16px;padding:16px 18px;
    box-shadow:var(--glass-shadow);transition:transform .15s ease;}
  .card:hover{transform:translateY(-2px);}
  .cl{font-size:11px;color:var(--text-secondary);text-transform:uppercase;letter-spacing:.5px;}
  .cv{font-size:24px;font-weight:700;margin:4px 0 2px;color:var(--text-primary);}
  .cs{font-size:11px;color:var(--text-secondary);}

  /* ── 图表面板 ── */
  .charts{display:grid;grid-template-columns:1fr 1fr;gap:14px;}
  .panel{background:var(--glass-bg);backdrop-filter:blur(18px);
    border:1px solid var(--glass-border);border-radius:16px;padding:16px 18px;
    box-shadow:var(--glass-shadow);}
  .panel h3{font-size:14px;margin:0 0 12px;font-weight:600;color:var(--text-primary);}
  .panel h3 .spark{color:var(--accent-light);margin-right:5px;font-size:13px;}
  .full{grid-column:1 / -1;}

  /* ── 表格 ── */
  table{width:100%;border-collapse:collapse;font-size:12px;}
  th,td{text-align:left;padding:7px 10px;border-bottom:1px solid rgba(180,200,190,.18);}
  th{color:var(--text-secondary);font-weight:500;font-size:11px;text-transform:uppercase;letter-spacing:.3px;
    background:rgba(180,200,190,.08);}
  td.num{text-align:right;font-variant-numeric:tabular-nums;font-feature-settings:'tnum';}
  .rpt-table td,.rpt-table th{padding:8px 12px;border-bottom:1px solid rgba(180,200,190,.22);}
  .rpt-table tbody tr:hover{background:rgba(168,212,188,.10);}

  /* ── 分布条形图（CSS） ── */
  .dist-row{display:flex;align-items:center;gap:10px;padding:5px 0;}
  .dist-label{width:36px;font-size:12px;color:var(--text-secondary);font-weight:500;text-align:last;}
  .dist-bar-wrap{flex:1;height:20px;background:rgba(180,200,190,.15);border-radius:10px;overflow:hidden;}
  .dist-bar{height:100%;border-radius:10px;background:linear-gradient(90deg,var(--accent-light), var(--accent));
    transition:width .4s ease;}
  .wd-bar{background:linear-gradient(90deg,#B8CDE0, #8FAEC8)!important;}
  .dist-val{width:90px;font-size:11px;color:var(--text-secondary);text-align:right;font-variant-numeric:tabular-nums;}

  /* ── 双列分布并排 ── */
  .dist-grid{display:grid;grid-template-columns:1fr 1fr;gap:20px;}
  .dist-grid .panel h3{font-size:13px;}

  /* ── 热词 chip ── */
  .kw-chips{display:flex;flex-wrap:wrap;gap:7px;margin-top:4px;}
  .chip{background:rgba(168,212,188,.20);border:1px solid rgba(168,212,188,.35);color:var(--text-primary);
    border-radius:999px;padding:5px 14px;font-size:12px;transition:background .15s;}
  .chip:hover{background:rgba(168,212,188,.35);}
  .chip b{color:var(--accent);margin-left:4px;}

  /* ── 上周订单 ── */
  .wk-sum{display:flex;flex-wrap:wrap;gap:7px;margin-bottom:10px;}
  .wk-chip{background:rgba(168,212,188,.15);border:1px solid rgba(168,212,188,.30);
    color:var(--text-primary);border-radius:999px;padding:4px 12px;font-size:12px;}
  .wk-chip.canc{background:var(--danger-bg);border-color:rgba(196,112,112,.25);color:var(--danger-text);}
  .wk-chip.ok{background:var(--ok-bg);border-color:rgba(122,170,144,.25);color:var(--ok-text);font-weight:600;}
  .wk-table-wrap{max-height:360px;overflow:auto;border-radius:12px;
    border:1px solid var(--glass-border);}
  .wk-table-wrap table th{position:sticky;top:0;background:var(--glass-bg);backdrop-filter:blur(14px);z-index:1;}
  .st-ok{color:var(--ok-text);font-weight:600;}
  .st-pend{color:var(--warn-text);font-weight:600;}
  .st-canc{color:var(--danger-text);font-weight:600;}

  /* ── 三列表格区 ── */
  .tables{display:grid;grid-template-columns:1fr 1fr 1fr;gap:14px;}
  .tables .panel{padding:14px 16px;}

  /* ── 报表导出提示 ── */
  .export-note{font-size:12px;color:var(--text-secondary);
    background:rgba(168,212,188,.10);border:1px solid rgba(168,212,188,.20);
    border-radius:10px;padding:10px 14px;margin-top:10px;}
  .export-note a{color:var(--accent);text-decoration:none;font-weight:500;}

  .note{font-size:12px;color:var(--text-secondary);margin-top:8px;}

  /* ── 响应式 ── */
  @media(max-width:900px){
    .cards{grid-template-columns:repeat(2,1fr);}
    .charts{grid-template-columns:1fr;}
    .tables{grid-template-columns:1fr;}
    .dist-grid{grid-template-columns:1fr;}
  }
</style></head><body><div class="wrap">

<!-- ═══ Header ═══ -->
<header>
  <h1><span class="spark">✦</span>睡sleepp的店 · 经营看板<span class="spark">✦</span></h1>
  <div class="meta">@PERIOD@ ｜ 生成 @GEN@ ｜ @FRESHWORD@</div>
  <div class="wow-badge">@WOW@</div>
</header>
@FRESH@

<!-- ═══ Section 1: 宏观大盘 ═══ -->
<div class="sec">
  <div class="sec-title"><span class="sec-num">1</span><span class="sec-name">宏观大盘</span></div>
  <div class="cards">@CARDS@</div>
  <div class="charts" style="margin-top:14px">
    <div class="panel"><h3><span class="spark">✦</span>月度趋势（净GMV + 订单）</h3><div style="position:relative;height:230px"><canvas id="mChart"></canvas></div></div>
    <div class="panel"><h3><span class="spark">✦</span>周度趋势（净GMV·已剔除取消）</h3><div style="position:relative;height:230px"><canvas id="wChart"></canvas></div></div>
  </div>
</div>

<!-- ═══ Section 2: 订单结构 ═══ -->
<div class="sec">
  <div class="sec-title"><span class="sec-num">2</span><span class="sec-name">订单结构</span></div>
  <div class="charts">
    <div class="panel"><h3><span class="spark">✦</span>各 KPOP 团体净 GMV</h3><div style="position:relative;height:230px"><canvas id="gChart"></canvas></div></div>
    <div class="panel"><h3><span class="spark">✦</span>经营健康度</h3><div style="position:relative;height:230px"><canvas id="hChart"></canvas></div>
      <div class="note">取消率/售后率越低越好，复购率越高越好。</div></div>
  </div>
  <div class="panel full" style="margin-top:14px"><h3><span class="spark">✦</span>现货 vs 预售 对比</h3>@SPOTPRE@
    <div class="note">预售取消率显著高于现货，是压制 GMV 的主要漏点。</div></div>
</div>

<!-- ═══ Section 3: 固定报表输出 ═══ -->
<div class="sec">
  <div class="sec-title"><span class="sec-num">3</span><span class="sec-name">固定报表输出</span></div>
  <div class="panel full"><h3><span class="spark">✦</span>月度汇总报表（可选中复制到 Excel）</h3>@MONTH_RPT@
    <div class="export-note">💡 提示：选中表格区域 → Ctrl+C → 在 Excel 中 Ctrl+V 即可粘贴。同时已导出 <a href="#">report_*.xlsx</a> 文件。</div></div>
  <div class="panel full" style="margin-top:14px"><h3><span class="spark">✦</span>季度汇总报表</h3>@QUARTER_RPT@
    <div class="export-note">按自然季度聚合，口径同月度报表。</div></div>
  <div class="dist-grid" style="margin-top:14px">
    <div class="panel"><h3>🕐 下单时段分布</h3>@TIME_DIST@</div>
    <div class="panel"><h3>📅 下单周几分布</h3>@WD_DIST@</div>
  </div>
</div>

<!-- ═══ Section 4: 销售热点 ═══ -->
<div class="sec">
  <div class="sec-title"><span class="sec-num">4</span><span class="sec-name">销售热点</span></div>
  <div class="panel full"><h3><span class="spark">✦</span>本周热销 SKU 关键词（@WEEKLBL@）</h3>@WEEKKW@
    <div class="note">基于最近 7 天已支付订单提取，关键词 = 团体 + 品类。</div></div>
  <div class="panel full" style="margin-top:14px"><h3><span class="spark">✦</span>上周具体订单（@WEEKLBL2@）</h3>
    <div class="wk-sum">@WEEKSUM@</div>
    <div class="wk-table-wrap">@WEEKTBL@</div>
    <div class="note">列出最近 7 天全部订单（含已取消），按金额降序。绿=完成，橙=处理中，红=已取消。</div></div>
  <div class="tables" style="margin-top:14px">
    <div class="panel"><h3><span class="spark">✦</span>TOP 商品</h3>@SKU@</div>
    <div class="panel"><h3><span class="spark">✦</span>地域分布</h3>@PROV@</div>
    <div class="panel"><h3><span class="spark">✦</span>团体对比</h3>@GROUP@</div>
  </div>
</div>

<div class="note" style="margin-top:20px;text-align:center">支付：@PAY@ ｜ 快递：@EXP@</div>

</div>
<script>const R=@JSON@;</script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.js"></script>
<script>
const PALETTE=['#A8D5BA','#F4B6C2','#C3B1E1','#A0C4E2','#F6E2A0','#F2B5A0','#B5C99A','#D6CDEA','#9FD8C8','#E8B4B8'];
function combo(id,labels,bar,line,bll,lll,bc,lc){
  new Chart(document.getElementById(id),{type:'bar',data:{labels:labels,datasets:[
    {type:'bar',label:bll,data:bar,backgroundColor:bc,borderRadius:6,yAxisID:'y',order:2},
    {type:'line',label:lll,data:line,borderColor:lc,backgroundColor:lc,borderWidth:2,tension:.3,pointRadius:3,yAxisID:'y1',order:1}]},
    options:{responsive:true,maintainAspectRatio:false,
      plugins:{legend:{labels:{font:{size:11},color:'#6B8076'}}},
      scales:{x:{ticks:{color:'#6B8076',font:{size:11}},grid:{display:false}},
        y:{position:'left',ticks:{color:'#6B8076'},grid:{color:'rgba(150,180,165,.10)'}},
        y1:{position:'right',ticks:{color:'#6B8076'},grid:{display:false}}}}});}
combo('mChart',R.月度.labels,R.月度.gmv,R.月度.orders,'净GMV','订单','#A8D5BA','#D4A0A0');
combo('wChart',R.周度.labels,R.周度.gmv,R.周度.orders,'周GMV','周订单','#A0C4E2','#C3B1E1');
new Chart(document.getElementById('gChart'),{type:'bar',
  data:{labels:R.团体.map(x=>x.name),datasets:[{label:'净GMV',data:R.团体.map(x=>x.gmv),
    backgroundColor:R.团体.map((x,i)=>PALETTE[i%PALETTE.length]),borderRadius:6}]},
  options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false},
    tooltip:{callbacks:{label:c=>'¥'+c.parsed.y}}},
    scales:{x:{ticks:{color:'#6B8076',font:{size:11}},grid:{display:false}},
      y:{ticks:{color:'#6B8076'},grid:{color:'rgba(150,180,165,.10)'}}}}});
new Chart(document.getElementById('hChart'),{type:'bar',data:{labels:['取消率','售后率','复购率'],
  datasets:[{label:'%',data:[R.取消率,R.售后率,R.复购率],
    backgroundColor:['#C47070','#D4A050','#7AAA90'],borderRadius:6,barThickness:32}]},
  options:{indexAxis:'y',responsive:true,maintainAspectRatio:false,
    plugins:{legend:{display:false},tooltip:{callbacks:{label:c=>c.parsed.x+'%'}}},
    scales:{x:{max:40,ticks:{color:'#6B8076'},grid:{color:'rgba(150,180,165,.10)'}},
      y:{ticks:{color:'#2D3E36',font:{size:13}},grid:{display:false}}}}});
</script></body></html>"""

    rep = {
        "@PERIOD@": R["数据周期"], "@GEN@": R["生成时间"], "@FRESHWORD@": freshword,
        "@WOW@": wow, "@FRESH@": fresh, "@CARDS@": cards_html,
        "@SKU@": sku_table(), "@PROV@": prov_table(), "@GROUP@": group_table(),
        "@SPOTPRE@": spotpre_table(),
        "@WEEKLBL@": R["本周标签"], "@WEEKKW@": week_hot_panel(),
        "@WEEKLBL2@": R["上周标签"], "@WEEKSUM@": wk_sum_html, "@WEEKTBL@": wk_tbl_html,
        "@PAY@": pay_str, "@EXP@": exp_str,
        "@MONTH_RPT@": month_report_table(), "@QUARTER_RPT@": quarter_report_table(),
        "@TIME_DIST@": time_dist_bars(), "@WD_DIST@": weekday_dist_bars(),
        "@JSON@": json.dumps(R, ensure_ascii=False),
    }
    for k, v in rep.items():
        tpl = tpl.replace(k, v)
    return tpl


def build_md(R):
    L = []
    L.append("# 睡sleepp的店 · 每周经营看板\n")
    L.append("- 数据周期：%s" % R["数据周期"])
    L.append("- 生成时间：%s" % R["生成时间"])
    L.append("- 数据新鲜度：距今天数 %d（最近订单 %s）\n" % (R["距今天数"], R["最近订单日"]))

    L.append("## 一、核心指标\n")
    L.append("| 指标 | 数值 |")
    L.append("|---|---|")
    L.append("| 净GMV(剔除取消) | ¥%.0f |" % R["净GMV"])
    L.append("| 已完成GMV | ¥%.0f |" % R["已完成GMV"])
    L.append("| 下单GMV(含取消) | ¥%.0f |" % R["下单GMV含取消"])
    L.append("| 订单数 | %d |" % R["订单数"])
    L.append("| 付费用户数 | %d |" % R["用户数"])
    L.append("| 客单价 | ¥%.0f（中位 ¥%.0f）|" % (R["客单价"], R["客单价中位"]))
    L.append("| 取消率 | %.1f%% |" % R["取消率"])
    L.append("| 售后率 | %.1f%% |" % R["售后率"])
    L.append("| 复购率 | %.1f%% |" % R["复购率"])
    L.append("| 预售占比 | %.0f%% |\n" % R["预售占比"])

    L.append("## 二、订单结构\n")
    sp, pr = R["现货"], R["预售"]
    L.append("| 指标 | 现货 | 预售 |")
    L.append("|---|---|---|")
    L.append("| 订单数 | %d | %d |" % (sp["订单数"], pr["订单数"]))
    L.append("| 净GMV | ¥%d | ¥%d |" % (sp["GMV"], pr["GMV"]))
    L.append("| 客单价 | ¥%.0f | ¥%.0f |" % (sp["客单价"], pr["客单价"]))
    L.append("| 取消率 | %.1f%% | %.1f%% |\n" % (sp["取消率"], pr["取消率"]))
    L.append("**各 KPOP 团体净 GMV**\n")
    L.append("| # | 团体 | 净GMV | 订单 |")
    L.append("|---|---|---|---|")
    for i, r in enumerate(R["团体"], 1):
        L.append("| %d | %s | ¥%d | %d |" % (i, r["name"], r["gmv"], r["orders"]))
    L.append("\n## 三、固定报表\n")
    L.append("**月度汇总**\n")
    L.append("| 月份 | 订单 | GMV | 客单价 | 用户 | 取消% | 售后% | 复购% | 预售% |")
    L.append("|---|---|---|---|---|---|---|---|---|")
    for m in R["月度报表"]:
        L.append("| %s | %d | ¥%d | ¥%.0f | %d | %.1f | %.1f | %.1f | %.0f |"
                 % (m["月份"], m["订单数"], m["GMV"], m["客单价"], m["用户数"],
                    m["取消率"], m["售后率"], m["复购率"], m["预售占比"]))
    L.append("\n**季度汇总**\n")
    L.append("| 季度 | 订单 | GMV | 客单价 | 用户 | 取消% | 售后% | 复购% | 预售% |")
    L.append("|---|---|---|---|---|---|---|---|---|")
    for q in R["季度报表"]:
        L.append("| %s | %d | ¥%d | ¥%.0f | %d | %.1f | %.1f | %.1f | %.0f |"
                 % (q["季度"], q["订单数"], q["GMV"], q["客单价"], q["用户数"],
                    q["取消率"], q["售后率"], q["复购率"], q["预售占比"]))
    L.append("\n**下单时段分布**\n")
    for t in R["时段分布"]:
        L.append("- **%s**：%d 单 / ¥%d（占 %.1f%%）/ 客单价 ¥%.0f" % (t["时段"], t["订单数"], t["GMV"], t["占比"], t["客单价"]))
    L.append("\n**下单周几分布**\n")
    for w in R["周几分布"]:
        L.append("- **%s**：%d 单 / ¥%d（占 %.1f%%）/ 客单价 ¥%.0f" % (w["周几"], w["订单数"], w["GMV"], w["占比"], w["客单价"]))
    L.append("\n## 四、销售热点\n")
    L.append("**本周热销 SKU 关键词（%s）**\n" % R["本周标签"])
    for k in R["本周热词"]:
        L.append("- %s：%d" % (k["word"], k["count"]))
    L.append("\n**本周 TOP SKU**\n")
    for i, r in enumerate(R["本周TOP"], 1):
        L.append("%d. %s — ¥%d（%d 单）" % (i, r["name"], r["gmv"], r["orders"]))
    L.append("\n**上周具体订单（%s）**\n" % R["上周标签"])
    su = R["上周汇总"]
    L.append("- 共 %d 单（现货 %d / 预售 %d），已取消 %d，净GMV ¥%d" % (su["单数"], su["现货"], su["预售"], su["取消"], su["净GMV"]))
    L.append("\n**TOP 商品**\n")
    for i, r in enumerate(R["TOP_SKU"], 1):
        L.append("%d. %s — ¥%d（%d 单）" % (i, r["name"], r["gmv"], r["orders"]))
    L.append("\n**地域 TOP**\n")
    for i, r in enumerate(R["省份"], 1):
        L.append("%d. %s — ¥%d（%d 单 / %d 用户）" % (i, r["name"], r["gmv"], r["orders"], r["users"]))
    L.append("\n## 履约\n")
    L.append("- 支付：%s" % "、".join("%s %d" % (k, v) for k, v in R["支付方式"].items()))
    L.append("- 快递：%s" % "、".join("%s %d" % (k, v) for k, v in R["快递"].items()))
    return "\n".join(L)


def main():
    # ── 1. 数据源 ──
    CUM = os.path.join(WS, "累计订单数据.xlsx")
    if len(sys.argv) > 1 and os.path.exists(sys.argv[1]):
        src = sys.argv[1]
    elif os.path.exists(CUM):
        src = CUM
    else:
        files = [f for f in glob.glob(os.path.join(WS, "*.xlsx"))]
        src = max(files, key=os.path.getmtime) if files else None
    if not src:
        print("ERROR: 未找到订单 .xlsx")
        sys.exit(1)
    print("数据源:", os.path.basename(src))

    # ── 2. 读取 + 清洗 ──
    df = read_orders(src)
    raw_rows = len(df)
    df = df[df["订单号"].notna()].copy()
    clean_rows = len(df)
    for c in ["订单创建时间", "支付时间", "订单发货时间", "订单完成时间"]:
        df[c] = pd.to_datetime(df[c], errors="coerce")
    df["月份"] = df["订单创建时间"].dt.to_period("M").astype(str)
    df["周起始"] = df["订单创建时间"].dt.to_period("W").apply(lambda p: p.start_time)

    paid = df[~df["订单状态"].isin(["已取消", "已关闭"])].copy()
    done = df[df["订单状态"].isin(["已完成", "已签收"])].copy()

    def money(x): return round(float(x), 2)

    # ── 3. 核心指标 ──
    R = {}
    R["数据周期"] = "%s ~ %s" % (df["订单创建时间"].min().strftime("%Y-%m-%d"),
                                  df["订单创建时间"].max().strftime("%Y-%m-%d"))
    R["生成时间"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    R["净GMV"] = money(paid[amt_col].sum())
    R["已完成GMV"] = money(done[amt_col].sum())
    R["下单GMV含取消"] = money(df[amt_col].sum())
    R["订单数"] = int(paid["订单号"].nunique())
    R["总包裹行"] = int(clean_rows)
    R["用户数"] = int(paid["用户编号"].nunique())
    order_gmv = paid.groupby("订单号")[amt_col].sum()
    R["客单价"] = money(order_gmv.mean())
    R["客单价中位"] = money(order_gmv.median())
    R["取消率"] = round(100 * (df["订单状态"] == "已取消").sum() / raw_rows, 1)
    R["售后率"] = round(100 * (df["售后状态"] == "售后完成").sum() / raw_rows, 1)
    uc = df.groupby("用户编号")["订单号"].nunique()
    R["复购率"] = round(100 * (uc >= 2).sum() / len(uc), 1)
    R["人均订单"] = round(uc.mean(), 2)
    R["预售占比"] = round(100 * (paid["订单类型"] == "预售订单").sum() / len(paid), 1)
    last_order = df["订单创建时间"].max()
    R["距今天数"] = (datetime.now() - last_order).days
    R["最近订单日"] = last_order.strftime("%Y-%m-%d")

    # ── 4. 时间序列（月度/周度）──
    all_months = sorted(paid["月份"].unique())
    m = paid.groupby("月份").agg(订单数=("订单号", "nunique"), GMV=(amt_col, "sum"))
    R["月度"] = {"labels": all_months, "gmv": [round(m.loc[mo, "GMV"]) for mo in all_months],
                  "orders": [int(m.loc[mo, "订单数"]) for mo in all_months]}
    end = paid["周起始"].max()
    start = end - pd.Timedelta(weeks=11)
    idx = pd.date_range(start, end, freq="W-MON")
    wk = paid.groupby("周起始").agg(订单数=("订单号", "nunique"), GMV=(amt_col, "sum")).reindex(idx, fill_value=0)
    R["周度"] = {"labels": [d.strftime("%m-%d") for d in idx],
                  "gmv": [round(x) for x in wk["GMV"]],
                  "orders": [int(x) for x in wk["订单数"]]}
    wg = R["周度"]["gmv"]
    R["最近周GMV"] = wg[-1]
    R["上周GMV"] = wg[-2]
    R["周环比"] = round(100 * (wg[-1] - wg[-2]) / wg[-2], 1) if wg[-2] > 0 else None

    # ── 5. 现货 vs 预售 ──
    spot = paid[paid["订单类型"] == "现货订单"]
    pre = paid[paid["订单类型"] == "预售订单"]
    df_spot = df[df["订单类型"] == "现货订单"]
    df_pre = df[df["订单类型"] == "预售订单"]
    R["现货"] = {
        "订单数": int(spot["订单号"].nunique()),
        "GMV": int(round(spot[amt_col].sum())),
        "客单价": money(spot.groupby("订单号")[amt_col].sum().mean()),
        "取消率": round(100 * (df_spot["订单状态"] == "已取消").sum() / len(df_spot), 1) if len(df_spot) else 0,
    }
    R["预售"] = {
        "订单数": int(pre["订单号"].nunique()),
        "GMV": int(round(pre[amt_col].sum())),
        "客单价": money(pre.groupby("订单号")[amt_col].sum().mean()),
        "取消率": round(100 * (df_pre["订单状态"] == "已取消").sum() / len(df_pre), 1) if len(df_pre) else 0,
    }

    # ── 6. 各 KPOP 团体 GMV ──
    paid2 = paid.copy()
    paid2["团体"] = paid2["SKU名称"].apply(group_of)
    grp = paid2.groupby("团体").agg(GMV=(amt_col, "sum"), 订单数=("订单号", "nunique"),
                                    件数=("SKU件数", "sum")).sort_values("GMV", ascending=False)
    R["团体"] = [{"name": str(i), "gmv": round(r["GMV"]), "orders": int(r["订单数"])} for i, r in grp.iterrows()]

    # ── 7. 本周热销 + 上周具体订单 ──
    max_day = df["订单创建时间"].max().normalize()
    win_start = max_day - pd.Timedelta(days=6)
    win_end = max_day + pd.Timedelta(days=1)
    wk_paid = paid[(paid["订单创建时间"] >= win_start) & (paid["订单创建时间"] < win_end)].copy()
    wk_all = df[(df["订单创建时间"] >= win_start) & (df["订单创建时间"] < win_end)].copy()
    wk_top = wk_paid.groupby("SKU名称").agg(GMV=(amt_col, "sum"), 订单数=("订单号", "nunique")).sort_values("GMV", ascending=False).head(8)
    R["本周TOP"] = [{"name": str(i)[:34], "gmv": round(r["GMV"]), "orders": int(r["订单数"])} for i, r in wk_top.iterrows()]
    kw_count = Counter()
    for name in wk_paid["SKU名称"].dropna().astype(str):
        for w in extract_keywords(name):
            kw_count[w] += 1
    R["本周热词"] = [{"word": w, "count": c} for w, c in kw_count.most_common(12)]
    R["本周标签"] = "%s ~ %s" % (win_start.strftime("%m-%d"), max_day.strftime("%m-%d"))
    R["上周标签"] = "%s ~ %s" % (win_start.strftime("%m-%d"), max_day.strftime("%m-%d"))

    wk_all = wk_all.sort_values(amt_col, ascending=False)
    def _st_cls(s):
        return "st-canc" if s == "已取消" else ("st-ok" if s in ("已完成", "已签收") else "st-pend")
    orows = ""
    for _, r in wk_all.iterrows():
        amt = round(float(r[amt_col])) if pd.notna(r[amt_col]) else 0
        orows += ("<tr><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td>%s</td><td class='num'>¥%d</td><td class='%s'>%s</td></tr>"
                 % (str(r["订单号"])[-8:],
                    (r["订单创建时间"].strftime("%m-%d") if pd.notna(r["订单创建时间"]) else ""),
                    str(r["SKU名称"])[:26], group_of(r["SKU名称"]),
                    str(r["订单类型"]).replace("订单", ""), amt,
                    _st_cls(str(r["订单状态"])), str(r["订单状态"])))
    R["上周订单html"] = ('<table><thead><tr><th>订单号</th><th>日期</th><th>SKU</th><th>团体>'
                         '<th>类型</th><th>金额</th><th>状态</th></tr></thead><tbody>%s</tbody></table>' % orows)
    su = {"单数": int(len(wk_all)),
          "现货": int((wk_all["订单类型"] == "现货订单").sum()),
          "预售": int((wk_all["订单类型"] == "预售订单").sum()),
          "取消": int((wk_all["订单状态"] == "已取消").sum()),
          "净GMV": int(round(wk_paid[amt_col].sum()))}
    R["上周汇总html"] = ('<span class="wk-chip">共 %d 单</span><span class="wk-chip">现货 %d</span>'
                         '<span class="wk-chip">预售 %d</span><span class="wk-chip canc">已取消 %d</span>'
                         '<span class="wk-chip ok">净GMV ¥%d</span>'
                         % (su["单数"], su["现货"], su["预售"], su["取消"], su["净GMV"]))
    R["上周汇总"] = su
    R["上周订单"] = [{"oid": str(r["订单号"])[-8:],
                      "日期": (r["订单创建时间"].strftime("%m-%d") if pd.notna(r["订单创建时间"]) else ""),
                      "sku": str(r["SKU名称"])[:30], "团体": group_of(r["SKU名称"]),
                      "类型": str(r["订单类型"]).replace("订单", ""),
                      "金额": int(round(float(r[amt_col]))) if pd.notna(r[amt_col]) else 0,
                      "状态": str(r["订单状态"])} for _, r in wk_all.iterrows()]

    # ── 8. 商品 / 地域 / 支付 / 快递 ──
    top_sku = paid.groupby("SKU名称").agg(件数=("SKU件数", "sum"), GMV=(amt_col, "sum"),
                                          订单数=("订单号", "nunique")).sort_values("GMV", ascending=False).head(10)
    R["TOP_SKU"] = [{"name": str(i)[:30], "gmv": round(r["GMV"]), "orders": int(r["订单数"])} for i, r in top_sku.iterrows()]
    prov = paid.groupby("省").agg(GMV=(amt_col, "sum"), 订单数=("订单号", "nunique"),
                                  用户数=("用户编号", "nunique")).sort_values("GMV", ascending=False).head(8)
    R["省份"] = [{"name": str(i), "gmv": round(r["GMV"]), "orders": int(r["订单数"]), "users": int(r["用户数"])} for i, r in prov.iterrows()]
    R["支付方式"] = {str(k): int(v) for k, v in paid["支付方式"].value_counts().items()}
    R["快递"] = {str(k): int(v) for k, v in paid["快递公司"].value_counts(dropna=False).head(6).items()}

    # ── 9. 【新增】下单时段 & 周几分布 ──
    paid_time = paid[paid["订单创建时间"].notna()].copy()
    paid_time["小时"] = paid_time["订单创建时间"].dt.hour
    paid_time["周几"] = paid_time["订单创建时间"].dt.weekday  # 0=Mon
    paid_time["订单GMV"] = paid_time[amt_col]

    # 时段统计
    time_stats = []
    total_gmv_t = paid_time[amt_col].sum()
    for pname, h_start, h_end in TIME_PERIODS:
        if h_end < h_start:  # 深夜 23-23 只取 23 点
            mask = paid_time["小时"] == h_start
        else:
            mask = (paid_time["小时"] >= h_start) & (paid_time["小时"] <= h_end)
        seg = paid_time[mask]
        seg_gmv = seg[amt_col].sum()
        seg_orders = seg["订单号"].nunique()
        seg_aov = seg.groupby("订单号")[amt_col].sum().mean() if seg_orders > 0 else 0
        time_stats.append({"时段": pname, "订单数": int(seg_orders), "GMV": int(round(seg_gmv)),
                           "客单价": round(seg_aov), "占比": round(100 * seg_gmv / total_gmv_t, 1) if total_gmv_t > 0 else 0})
    R["时段分布"] = time_stats

    # 周几统计
    wd_stats = []
    for di, dname in enumerate(WEEKDAY_NAMES):
        dw = paid_time[paid_time["周几"] == di]
        dw_gmv = dw[amt_col].sum()
        dw_orders = dw["订单号"].nunique()
        dw_aov = dw.groupby("订单号")[amt_col].sum().mean() if dw_orders > 0 else 0
        wd_stats.append({"周几": dname, "订单数": int(dw_orders), "GMV": int(round(dw_gmv)),
                         "客单价": round(dw_aov), "占比": round(100 * dw_gmv / total_gmv_t, 1) if total_gmv_t > 0 else 0})
    R["周几分布"] = wd_stats

    # ── 10. 【新增】月度 / 季度报表 ──
    # 月度报表（每行一个完整月的核心指标）
    monthly_rpt = []
    for mo in all_months:
        mo_paid = paid[paid["月份"] == mo]
        mo_all = df[df["月份"] == mo]
        mo_raw = len(mo_all)
        mo_done = mo_all[mo_all["订单状态"].isin(["已完成", "已签收"])]
        mo_uc = mo_all.groupby("用户编号")["订单号"].nunique()
        mo_orders = mo_paid["订单号"].nunique()
        mo_gmv = mo_paid[amt_col].sum()
        mo_aov = mo_paid.groupby("订单号")[amt_col].sum().mean() if mo_orders > 0 else 0
        mo_cancel_rate = round(100 * (mo_all["订单状态"] == "已取消").sum() / mo_raw, 1) if mo_raw > 0 else 0
        mo_after_rate = round(100 * (mo_all["售后状态"] == "售后完成").sum() / mo_raw, 1) if mo_raw > 0 else 0
        mo_rebuy_rate = round(100 * (mo_uc >= 2).sum() / len(mo_uc), 1) if len(mo_uc) > 0 else 0
        mo_pre_pct = round(100 * (mo_paid["订单类型"] == "预售订单").sum() / len(mo_paid), 1) if len(mo_paid) > 0 else 0
        monthly_rpt.append({
            "月份": mo, "订单数": int(mo_orders), "GMV": int(round(mo_gmv)),
            "客单价": round(mo_aov), "用户数": int(mo_paid["用户编号"].nunique()),
            "取消率": mo_cancel_rate, "售后率": mo_after_rate,
            "复购率": mo_rebuy_rate, "预售占比": mo_pre_pct,
        })
    R["月度报表"] = monthly_rpt

    # 季度报表
    def _quarter_label(mo_str):
        y, m = mo_str.split("-")
        m = int(m)
        q = (m - 1) // 3 + 1
        return "%sQ%d" % (y, q)

    paid_copy = paid.copy()
    paid_copy["季度"] = paid_copy["月份"].apply(_quarter_label)
    df_copy = df.copy()
    df_copy["季度"] = df_copy["月份"].apply(_quarter_label)
    quarters = sorted(paid_copy["季度"].unique())
    quarterly_rpt = []
    for q in quarters:
        q_paid = paid_copy[paid_copy["季度"] == q]
        q_all = df_copy[df_copy["季度"] == q]
        q_raw = len(q_all)
        q_uc = q_all.groupby("用户编号")["订单号"].nunique()
        q_orders = q_paid["订单号"].nunique()
        q_gmv = q_paid[amt_col].sum()
        q_aov = q_paid.groupby("订单号")[amt_col].sum().mean() if q_orders > 0 else 0
        q_cancel_rate = round(100 * (q_all["订单状态"] == "已取消").sum() / q_raw, 1) if q_raw > 0 else 0
        q_after_rate = round(100 * (q_all["售后状态"] == "售后完成").sum() / q_raw, 1) if q_raw > 0 else 0
        q_rebuy_rate = round(100 * (q_uc >= 2).sum() / len(q_uc), 1) if len(q_uc) > 0 else 0
        q_pre_pct = round(100 * (q_paid["订单类型"] == "预售订单").sum() / len(q_paid), 1) if len(q_paid) > 0 else 0
        quarterly_rpt.append({
            "季度": q, "订单数": int(q_orders), "GMV": int(round(q_gmv)),
            "客单价": round(q_aov), "用户数": int(q_paid["用户编号"].nunique()),
            "取消率": q_cancel_rate, "售后率": q_after_rate,
            "复购率": q_rebuy_rate, "预售占比": q_pre_pct,
        })
    R["季度报表"] = quarterly_rpt

    # ── 11. 导出 ──
    with open(os.path.join(WS, "dashboard_data.json"), "w", encoding="utf-8") as f:
        json.dump(R, f, ensure_ascii=False, indent=2)
    df.to_csv(os.path.join(WS, "clean_orders.csv"), index=False, encoding="utf-8-sig")

    # 导出 Excel 报表
    excel_path = _export_excel(R, paid, df)

    # ── 12. 渲染 ──
    with open(os.path.join(WS, "weekly_dashboard.html"), "w", encoding="utf-8") as f:
        f.write(build_html(R))
    with open(os.path.join(WS, "weekly_report.md"), "w", encoding="utf-8") as f:
        f.write(build_md(R))

    print("OK -> weekly_dashboard.html / weekly_report.md / clean_orders.csv / dashboard_data.json")
    print("Excel ->", os.path.basename(excel_path))
    print("净GMV=%.0f 订单=%d 用户=%d 取消率=%.1f%% 售后率=%.1f%% 复购率=%.1f%%"
          % (R["净GMV"], R["订单数"], R["用户数"], R["取消率"], R["售后率"], R["复购率"]))
    print("时段高峰:", max(R["时段分布"], key=lambda x: x["GMV"])["时段"],
          "| 周几高峰:", max(R["周几分布"], key=lambda x: x["GMV"])["周几"])


if __name__ == "__main__":
    main()
